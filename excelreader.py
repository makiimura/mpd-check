import openpyxl


class InputReader(object):

    def __init__(self, inputFile):
        self.data = list()
        workbook = openpyxl.load_workbook(inputFile)
        sheet = workbook.get_sheet_by_name('videoes');

        for i in range(2, sheet.max_row):
            item = {'ContentID' : int(sheet.cell(row=i, column=1).value),
                    'Profile' : str(sheet.cell(row=i, column=2).value),
                    'URL': str(sheet.cell(row=i, column=3).value)}

            self.data.append(item)

    def hasNext(self):
        return (len(self.data) != 0)

    def fetchRow(self):
        return self.data.pop(0)


def test():
    tf = r'./test.xlsx'
    reader = InputReader(tf)

    while reader.hasNext():
        row = reader.fetchRow()

        # Get profile ID
        print "-"*100
        print "Content ID: %d" % row['ContentID']
        print "Profile: %s" % row['Profile']
        print "URL: %s" % row['URL']


if __name__=='__main__':
    test()
